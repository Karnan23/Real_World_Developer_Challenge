import logging, os, datetime, json,smtplib
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

try:
    load_dotenv()
    sender_mail = os.getenv("SENDER_EMAIL")
    app_password = os.getenv("APP_PASSWORD")
    receiver_mail = os.getenv("RECEIVER_EMAIL")
    report_name = os.getenv("REPORT_NAME")
    email_subject = os.getenv("EMAIL_SUBJECT")
    smtp_server=os.getenv("SMTP_SERVER")
    smtp_port=int(os.getenv("SMTP_PORT"))
    news_path = os.getenv("NEWS_PATH")

except Exception as e:
    logging.error(f"Error loading configuration: {e}")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',handlers=[logging.FileHandler("report.log"), logging.StreamHandler()])
logging.info("Configuration loaded successfully.")

logging.info("Script Started...")

def load_news_data(news_path, no_of_news):
    try:
        if not os.path.exists(news_path):
            logging.error(f"News file not found: {news_path}")
            return []
        titles = []
        urls = []
        authors = []
        all_published_at = []
        sources = []
        with open(news_path, 'r') as file:
            news_data = json.load(file)
        for data in news_data[:int(no_of_news)]:
            title = data.get("title", "No Title")
            url = data.get("url", "No URL")
            author = data.get("author", "No Author")
            published_at = data.get("published_at", "No Published Date")
            source = data.get("source", "No source")
            logging.info(f"Loaded news article: {title}")

            titles.append(title)
            urls.append(url)
            authors.append(author)
            all_published_at.append(published_at)
            sources.append(source)
        logging.info("news datas are organized successfully.")

        return [{"title": t, "url": u, "author": a, "published_at": p, "source": s} for t, u, a, p, s in zip(titles, urls, authors, all_published_at, sources)]

    except Exception as e:
        logging.error(f"Error loading news data: {e}")
        return []


def create_excel_report(news_datas, report_name):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Tech Trends Report"
        headers = ["Title", "URL", "Author", "Published At", "Source"]
        ws.append(headers)
       
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        logging.info("Excel report headers created successfully.")

        for row_num, item in enumerate(news_datas, start=2):
            ws.cell(row=row_num, column=1, value=item["title"])
            ws.cell(row=row_num, column=2, value=item["url"])
            ws.cell(row=row_num, column=3, value=item["author"])
            ws.cell(row=row_num, column=4, value=item["published_at"])
            ws.cell(row=row_num, column=5, value=item["source"])
            logging.info(f"Added data row {row_num} successfully.")

        for col_idx,col in enumerate(ws.columns,1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        logging.info("Excel report formatted successfully.")
        wb.save(report_name)
        logging.info(f"Excel report file created successfully: {report_name}")

    except Exception as e:
        logging.error(f"Error creating excel report: {e}")

def send_mail(report_name):
    try:
        report_path=os.path.abspath(report_name)
        msg= MIMEMultipart()
        msg['From']=f"TECH NEWS REPORT <{sender_mail}>"
        msg['To']=receiver_mail
        msg['Subject']=f"{email_subject}-{datetime.datetime.now():%Y-%m-%d}"

        attachment=open(report_path,'rb')
        part=MIMEBase('application','octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % report_name)
        msg.attach(part)
        attachment.close()
        logging.info("Email msg created successfully")


        with smtplib.SMTP(smtp_server,smtp_port) as mail:
            mail.starttls()
            mail.login(sender_mail,app_password)
            mail.send_message(msg)
            logging.info("Email send successfully!")
    except Exception as e:
        logging.error(f"Email error : {e}")

if __name__ == "__main__":
    try:
        news_datas = load_news_data(news_path, 3)
        create_excel_report(news_datas, report_name)
        send_mail(report_name)
        logging.info("Script ended and completed successfully...")
    except Exception as e:
        logging.error(f"Script run error: {e}")
